import re
import sys
from pathlib import Path

import openpyxl
import pyodbc


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python import_excel_to_sql.py <excel_path>")

    excel_path = Path(sys.argv[1])
    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    raw_headers = [c.value for c in ws[1]]
    headers = []
    seen = {}
    for i, h in enumerate(raw_headers, start=1):
        name = str(h).strip() if h is not None else ""
        if not name:
            name = f"Column_{i}"
        name = re.sub(r"\s+", " ", name)
        base = name
        n = seen.get(base, 0) + 1
        seen[base] = n
        if n > 1:
            name = f"{base}_{n}"
        headers.append(name)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append([None if v is None else str(v) for v in row])

    conn = pyodbc.connect(
        "DRIVER={ODBC Driver 18 for SQL Server};SERVER=.;Trusted_Connection=yes;Encrypt=no;",
        autocommit=True,
    )
    cur = conn.cursor()
    cur.execute("IF DB_ID('ExcelTtbDB') IS NULL CREATE DATABASE [ExcelTtbDB];")
    conn.close()

    conn = pyodbc.connect(
        "DRIVER={ODBC Driver 18 for SQL Server};SERVER=.;DATABASE=ExcelTtbDB;Trusted_Connection=yes;Encrypt=no;"
    )
    cur = conn.cursor()

    table_name = "OCR_TTB_17"
    cur.execute(f"IF OBJECT_ID('dbo.{table_name}','U') IS NOT NULL DROP TABLE dbo.{table_name};")

    col_defs = ", ".join([f"[{h.replace(']', ']]')}] NVARCHAR(MAX) NULL" for h in headers])
    cur.execute(f"CREATE TABLE dbo.{table_name} ({col_defs});")

    placeholders = ", ".join(["?"] * len(headers))
    col_list = ", ".join([f"[{h.replace(']', ']]')}]" for h in headers])
    insert_sql = f"INSERT INTO dbo.{table_name} ({col_list}) VALUES ({placeholders})"
    cur.fast_executemany = True
    cur.executemany(insert_sql, rows)
    conn.commit()

    cur.execute(f"SELECT COUNT(*) FROM dbo.{table_name};")
    row_count = cur.fetchone()[0]
    cur.execute(
        f"SELECT COUNT(*) FROM sys.columns WHERE object_id = OBJECT_ID('dbo.{table_name}');"
    )
    col_count = cur.fetchone()[0]

    print(f"DB=ExcelTtbDB TABLE=dbo.{table_name} ROWS={row_count} COLS={col_count}")
    conn.close()


if __name__ == "__main__":
    main()
